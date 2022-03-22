# -*- coding: utf-8 -*-
"""
Created on Tue Feb 15 17:58:03 2022

@author: wroser
"""

import os
import shutil
import datetime
import openpyxl as xl

def make_output_directory(name=None):
    if not name is None:
        filename = name
    else:
        filename = input('Please name the output folder. '
                         'Press enter to make a temporary folder. '
                         'Enter "time" to use the current date and time.\n')
        if filename == '':
            filename = 'temp'
        elif filename in ['t', 'time']:
            now = datetime.datetime.now()
            filename = now.strftime('%Y-%m-%d_%H%M')
    try:
        os.mkdir('out')
    except FileExistsError:
        pass
    try:
        os.mkdir('out/' + filename)
    except FileExistsError:
        newfilename = input('Folder already exists. Please choose a different '
                            'name or press enter to overwrite the existing '
                            'temp folder.\n')
        if newfilename == '':
            shutil.rmtree('out/' + filename)
            os.mkdir('out/' + 'temp')
        else:
            try:
                os.mkdir('out/' + newfilename)
            except FileExistsError:
                shutil.rmtree('out/' + newfilename)
                os.mkdir('out/' + newfilename)
    return 'out/' + filename, filename

def save_input_file(name, output_directory):
    keep_these = ['nodes', 'nodeFix', 'nodeMass', 'diaphragms', 'elements', 'eleProperties', 'eleTransf']
    input_wb = xl.load_workbook(name, keep_vba=False, data_only=True)
    out_wb = xl.Workbook()
    for sheet in input_wb.sheetnames:
        if sheet in keep_these:
            input_wb[sheet]._parent = out_wb
            out_wb._add_sheet(input_wb[sheet])
        else:
            pass
    del out_wb['Sheet'] # This is a default sheet created by openpyxl.
    i_file_ext = -1*(len(name) - name.index('.'))
    out_wb.save('out/' + output_directory + '/' + name[:i_file_ext] + '.xlsx')
    out_wb.close()
    input_wb.close()

def _check_if_exists(file):
    pass

def _remove_output(folder):
    shutil.rmtree('out/' + folder)